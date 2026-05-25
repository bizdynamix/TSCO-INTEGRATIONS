#!/usr/bin/env python3
"""Generate Excel audit sheet for migrated Language Profiles"""

import requests
import os
import json
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()

MONDAY_API_TOKEN = os.getenv('MONDAY_API_TOKEN')
BOARD_ID = "8445103301"

def get_monday_data():
    """Fetch all items with Language Profile files from Monday.com"""
    headers = {
        'Authorization': MONDAY_API_TOKEN,
        'Content-Type': 'application/json'
    }
    
    all_items = []
    cursor = None
    
    while True:
        if cursor:
            query = f'''{{
                next_items_page(limit: 500, cursor: "{cursor}") {{
                    cursor
                    items {{
                        id
                        name
                        column_values {{
                            id
                            value
                        }}
                        assets {{
                            id
                            name
                        }}
                    }}
                }}
            }}'''
        else:
            query = f'''{{
                boards(ids: {BOARD_ID}) {{
                    items_page(limit: 500) {{
                        cursor
                        items {{
                            id
                            name
                            column_values {{
                                id
                                value
                            }}
                            assets {{
                                id
                                name
                            }}
                        }}
                    }}
                }}
            }}'''
        
        response = requests.post(
            'https://api.monday.com/v2',
            json={'query': query},
            headers=headers
        )
        data = response.json()
        
        if cursor:
            items_page = data['data']['next_items_page']
        else:
            items_page = data['data']['boards'][0]['items_page']
        
        all_items.extend(items_page['items'])
        print(f"   Fetched {len(all_items)} items...")
        
        cursor = items_page.get('cursor')
        if not cursor:
            break
    
    return all_items

def create_audit_sheet():
    """Create Excel audit sheet with migrated profiles"""
    print("📥 Fetching data from Monday.com...")
    items = get_monday_data()
    
    # Process items - filter for ACTIVE with valid Language Profile
    audit_data = []
    
    # Status index 6 = "Active" based on column settings
    ACTIVE_INDEX = 6
    
    for item in items:
        cols = {c['id']: c['value'] for c in item['column_values']}
        
        # Get Lang Status (uses index, not label)
        status_val = cols.get('status', '{}')
        try:
            status_data = json.loads(status_val) if status_val else {}
            status_index = status_data.get('index') if isinstance(status_data, dict) else None
        except:
            status_index = None
        
        if status_index != ACTIVE_INDEX:
            continue
        
        # Get Partner Project Name
        project_val = cols.get('project_mkm1qfap', '""')
        try:
            project_name = json.loads(project_val) if project_val else ''
            if not project_name:
                project_name = ''
        except:
            project_name = ''
        
        # Check if has valid Language Profile
        file_val = cols.get('files0', '{}')
        file_name = ''
        try:
            file_data = json.loads(file_val) if file_val else {}
            if 'files' not in file_data or not file_data['files']:
                continue
            
            file_name = file_data['files'][0].get('name', '')
            asset_id = file_data['files'][0].get('assetId')
            if not asset_id:
                continue
            
            # Verify asset exists
            asset_exists = any(str(a['id']) == str(asset_id) for a in item['assets'])
            if not asset_exists:
                continue
                
        except:
            continue
        
        # Get Partner Name (text7 column based on board structure)
        partner_val = cols.get('text7', '""')
        try:
            partner_name = json.loads(partner_val) if partner_val else ''
            if not partner_name:
                partner_name = ''
        except:
            partner_name = ''
        
        # Get Marketing Region if available
        region_val = cols.get('dropdown', '{}')  # Common column ID for dropdowns
        try:
            region_data = json.loads(region_val) if region_val else {}
            marketing_region = ', '.join(region_data.get('labels', [])) if isinstance(region_data, dict) else ''
        except:
            marketing_region = ''
        
        audit_data.append({
            'project_name': project_name.upper() if project_name else item['name'].upper(),
            'language': item['name'],
            'partner_name': partner_name,
            'file_name': file_name,
            'audit_notes': '',
            'marketing_region': marketing_region
        })
    
    # Sort by project name
    audit_data.sort(key=lambda x: x['project_name'])
    
    print(f"📊 Creating Excel sheet with {len(audit_data)} records...")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Language Profile Audit"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Project Name", "Language", "Partner Name", "Language Profile File", "Audit Notes", "Marketing Region"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row, record in enumerate(audit_data, 2):
        ws.cell(row=row, column=1, value=record['project_name']).border = thin_border
        ws.cell(row=row, column=2, value=record['language']).border = thin_border
        ws.cell(row=row, column=3, value=record['partner_name']).border = thin_border
        ws.cell(row=row, column=4, value=record['file_name']).border = thin_border
        ws.cell(row=row, column=5, value=record['audit_notes']).border = thin_border
        ws.cell(row=row, column=6, value=record['marketing_region']).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save
    filename = f"Language_Profile_Audit_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = f"/Users/edwinbrooks/Projects/MONDAY-SHAREPOINT/{filename}"
    wb.save(filepath)
    
    print(f"✅ Created: {filename}")
    print(f"   Records: {len(audit_data)}")
    
    return filepath

if __name__ == "__main__":
    create_audit_sheet()
